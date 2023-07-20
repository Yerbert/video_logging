#!/usr/bin/env python3

from fileinput import filename
import rospy
from std_msgs.msg import String
from video_logging.msg import FilterSwitch
import numpy as np
import time
from datetime import datetime
import random
import os
import subprocess
import signal
import pickle
import argparse
import rospkg
import copy
import sys
import signal
try:
    from builtins import input
except ImportError:
    from __builtin__ import input
#from Constants import Constants
from geometry_msgs.msg import Point, Vector3, WrenchStamped, PoseStamped
from openpyxl import load_workbook, Workbook

import rosbag_player_experimental
from JackalSSH import JackalSSH

# For printing formatting
class Color:
   PURPLE = '\033[95m'
   MAGENTA = '\033[35m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'

class Errors:
    name = None
    types = {
                        "A": "Motor Issue",
                        "B": "Camera Smudge",
                        "C": "LaserScanner Issue (Fake Obstruction)",
                        "D": "Localisation Issue",
                        "E": "Dropped Payload",
                        "F": "Camera Issue",
                        "G": "Obstruction in Robot's Path",
                        "H": "LaserScanner Flicker",
                        "T": "Training Scenario",
                        "R": "Regular"
    }
    rosbags = {
                        "A": "MotorIssue.bag",
                        "B": "CameraSmudge.bag",
                        "C": "FakeObstruction.bag",
                        "D": "LocalisationIssue.bag",
                        "E": "DroppedPayload.bag",
                        "F": "CameraFailure.bag",
                        "G": "GoalObstruction.bag",
                        "H": "VelodyneError.bag",
                        "T": "TrainingRecording.bag",
                        "R": "Regular.bag"
    }


"""
# Prior to 10/07/23
    types = {
                        "A": "Flat tyre",
                        "B": "Motor Failure",
                        "C": "Camera Sensor Failure",
                        "D": "Velodyne LIDAR Failure",
                        "E": "Obstacle in the Way",
                        "F": "Robot Senses Non-existent Object",
                        "G": "Dropped Payload",
                        "H": "Velodyne LIDAR Failure and Localisation Error",
                        "T": "Training Scenario",
                        "R": "Regular"
    }
    rosbags = {
                        "A": "FlatTyre.bag",
                        "B": "MotorFailure.bag",
                        "C": "CameraFailure.bag",
                        "D": "VelodyneError.bag",
                        "E": "GoalObstruction.bag",
                        "F": "FakeObstruction.bag",
                        "G": "DroppedPayload.bag",
                        "H": "LocalisationError.bag",
                        "T": "TrainingRecording.bag",
                        "R": "Regular.bag"
    }
"""

class Conditions:
    name = None
    conditions = {
                        "1": "AR + replay",
                        "2": "AR + live",
                        "3": "Tablet + replay",
                        "4": "Tablet + live"
                        }

class MyWorkbook:
    def __init__(self):
        self.conditionCol = 3
        self.errorCol = 4
        self.correctErrorCol = 5
        self.guessedErrorCol = 6
        self.diagnoseTimeCol = 7
        #ConfidenceTimeCol = 6
        self.dateCol = 1
        self.partIdCol = 2
        self.confidenceCol = 8
        self.questionCorrectCol = 9
        self.followupQn1TimeCol = 10
        self.followupQn2TimeCol = 11
    
        self.file_path = rospkg.RosPack().get_path('video_logging') + '/Sheets'
        self.excel_filename = "data.xlsx"
        self.fullpath = os.path.join(self.file_path, self.excel_filename)
        self.wb = load_workbook(self.fullpath)
        self.ws = self.wb.active
    
    def write_next_row(self, participant_no, condition, error, correct_error, guessed_error, diagnosis_time, confidence, question_correct, followup_qn_1_time, followup_qn_2_time):
        row = self.ws.max_row + 1
        self.ws.cell(row=row, column=self.dateCol).value = datetime.now()
        self.ws.cell(row=row, column=self.partIdCol).value = participant_no
        self.ws.cell(row=row, column=self.conditionCol).value = condition
        self.ws.cell(row=row, column=self.errorCol).value = error
        self.ws.cell(row=row, column=self.correctErrorCol).value = correct_error
        self.ws.cell(row=row, column=self.guessedErrorCol).value = guessed_error
        self.ws.cell(row=row, column=self.diagnoseTimeCol).value = diagnosis_time
        self.ws.cell(row=row, column=self.confidenceCol).value = confidence
        self.ws.cell(row=row, column=self.questionCorrectCol).value = question_correct
        self.ws.cell(row=row, column=self.followupQn1TimeCol).value = followup_qn_1_time
        self.ws.cell(row=row, column=self.followupQn2TimeCol).value = followup_qn_2_time
        self.wb.save(self.fullpath)


class AR_error_diagnostics:
    def __init__(self, participant_no):
        self.participant_no = int(participant_no)
        self.conditions = [""]*8
        self.errors = [""]*8
        self.condition_pub = rospy.Publisher("/condition", String, queue_size=10)

    def signal_handler(self, sig, frame):
        print("\n\nshutting down from AR_error_diagnostics\n\n")
        rospy.sleep(.5)
        sys.exit(0)

    def all_loop(self):

        # Block live data stream initially
        print("  Blocking all data streams initially...")
        filters = FilterSwitch(
            velodyne_blocked=True,
            camera_blocked=True
        )
        # JackalSSH().ros_pub_filterswitch(filters).kill()
        j1 = JackalSSH().ros_pub_msg("/filters", "video_logging/FilterSwitch", filters)
        # Turn fake objects and calibration features off
        print("  Disabling fake object...")
        j2 = JackalSSH().ros_pub("/fake_object", "std_msgs/Bool", "false")
        os.system("rostopic pub -1 /fake_object std_msgs/Bool false")
        print("  Disabling calibration tools...")
        os.system("rostopic pub -1 /calibration std_msgs/Bool false")
        
        # Kill ssh's
        j1.kill(4)
        j2.kill(0)

        # Signal handler
        signal.signal(signal.SIGINT, self.signal_handler)



        # TEST
        if self.participant_no == 0:
            self.conditions = ["1"]
            self.errors = ["R"]

        # Test all errors on a single condition
        elif self.participant_no < 0:
            self.conditions = [str(abs(self.participant_no))]*8
            self.errors = ["A","B","C","D","E","F","G","H"]
        
        # Regular User Study Participant
        else:
            #Find and open file with conditions and errors used by each participant
            pol_file_path = rospkg.RosPack().get_path('video_logging') + '/Sheets'
            pol_filename = "pol.xlsx"
            pol_wb = load_workbook(os.path.join(pol_file_path, pol_filename))
            pol_sheet = pol_wb.active
            pol_rows = 33
            pol_columns = pol_sheet.max_column
            participant_row = 0

            #Find row relating to current participant
            for i in range(2,pol_rows):
                number = int(pol_sheet.cell(row = i, column = 1).value)
                if (number%32) == self.participant_no:
                    participant_row = i
            
            #Find order of conditions for current participant
            j = 0
            for i in range(2,6):
                condition = int(pol_sheet.cell(row = participant_row, column = i).value)
                self.conditions[j] = str(condition)
                j = j + 1
                self.conditions[j] = str(condition)
                j = j + 1


            #Find order of errors for current participant
            for i in range(len(self.errors)):
                col = i+6
                self.errors[i] = pol_sheet.cell(row = participant_row, column = col).value
            
            #Prepend static training scenarios to lists
            self.errors = ["T","T","T","T"] + self.errors
            self.conditions = ["4","3","1","2"] + self.conditions


        #Display all scenarios
        print("\nThe order of scenarios will be:")
        for err,cond in zip(self.errors,self.conditions):
            print("  {0: <17}".format(Conditions.conditions[cond]), Errors.types[err])
        print("\n")
        # workbook = MyWorkbook()

        input("Run through first page of instructions if not already done. Press enter to continue once completed    ")

        #Go through each error
        for l, error in enumerate(self.errors):
            #Print error information here
            print("\n"+ str(max(0,l+1-4)) +". The error will be " + Color.BOLD + Color.CYAN + Errors.types[self.errors[l]] + Color.END)
            #Print condition for operator to know which method is being used
            print("The condition required for this error is " + Color.BOLD + Color.RED + Conditions.conditions[self.conditions[l]] + Color.END)
            cont = input("Do you want to perform this Error/condition? (Y/N)  ")
            check = 0
            if cont == "Y" or cont == "N":
                check = 1
            while check != 1:
                cont = input("Error. Invalid Input. Do you want to perform this Error/condition? (Y/N)  ")
                if cont == "Y" or cont == "N":
                    check = 1
            if cont == "N":
                print("skipping error " + Errors.types[self.errors[l]] + "\n\n\n")
                continue

            self.configure_device_connections(Conditions.conditions[self.conditions[l]], Errors.rosbags[self.errors[l]])

            wait = input("\nPrepare condition now and ensure devices are connected.\nPress any key to begin streaming data   ")
            print("\n")
            #Once user has confirmed ready, switch to other file to start condition
            
            
            play_condition = rosbag_player_experimental.Run_Condition()
            data_to_write = play_condition.run_condition(Errors.rosbags[self.errors[l]],Errors.types[self.errors[l]],Conditions.conditions[self.conditions[l]])

            # Record stats for non-training scenarios
            if l > 3 or self.participant_no == 0:
                # MyWorkbook().write_next_row(participant_no,Conditions.conditions[self.conditions[l]],Errors.types[self.errors[l]],data_to_write[0],data_to_write[1],data_to_write[2],data_to_write[3],data_to_write[4])
                MyWorkbook().write_next_row(participant_no,Conditions.conditions[self.conditions[l]],Errors.types[self.errors[l]], *data_to_write)
                print("Error " + str(max(0,l+1-4)) + " completed\n\n\n")
            
            # Prompt experimentor to get user to do post-condition survey
            if l > 4 and (self.conditions[l] == self.conditions[l-1]):
                wait = input("\n\nUSER NOW DOES POST-CONDITION SURVEY")
                print("\n")

            # End of training
            if l == 3:
                input("\nProvide instructions on difference between live and replay and then press enter    ")
                print("\n\nTraining complete. About to move onto study.... \n\n")
            
        print("All error conditions complete, exiting...\n\nUSER NOW DOES END-OF-STUDY SURVEY\n")
    

    def configure_device_connections(self, new_condition, rosbag_name):

        sleep_seconds = 4
        print("\n  Signalling devices to configure connections...")
        
        self.condition_pub.publish(String(new_condition))
        JackalSSH().ros_pub_condition(new_condition).kill(3)
        print("  Sleeping for {} more seconds to allow connections...".format(sleep_seconds))
        rospy.sleep(sleep_seconds) # to allow reconnections to occur

if __name__ == '__main__':
    rospy.init_node('error_diagnostics_user_study')
    participant_no = input("Enter Participant Number (0 = Test): ")
    
    AR_error_diagnostics_obj = AR_error_diagnostics(participant_no)
    AR_error_diagnostics_obj.all_loop() 
        # filename = 'participant_' + args.participant_ID + '.pkl'
        # pickle.dump(handover_obj.seq_list, open(filename, "wb"))
