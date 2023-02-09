#!/usr/bin/env python3

from fileinput import filename
import rospy
import numpy as np
import time
from datetime import datetime
import random
import os
import subprocess
import signal
import pickle
import argparse
import rospkg
import copy
import sys
import signal
try:
    from builtins import input
except ImportError:
    from __builtin__ import input
#from Constants import Constants
from geometry_msgs.msg import Point, Vector3, WrenchStamped, PoseStamped
from openpyxl import load_workbook, Workbook

import rosbag_player_experimental

class Errors:
    name = None
    types = {
                        "A": "Flat tyre",
                        "B": "Motor Failure",
                        "C": "Camera Sensor Failure",
                        "D": "Velodyne LIDAR Failure",
                        "E": "Obstacle in the Way",
                        "F": "Robot Senses Non-existent Object",
                        "G": "Dropped Payload",
                        "H": "Localisation Error"
    }
    rosbags = {
                        "A": "FlatTyre.bag",
                        "B": "MotorFailure.bag",
                        "C": "CameraFailure.bag",
                        "D": "VelodyneError.bag",
                        "E": "HumanObstruction.bag",
                        "F": "HumanObstruction.bag",
                        "G": "DroppedPayload.bag",
                        "H": "LocalisationError.bag"
                        }

class Conditions:
    name = None
    conditions = {
                        "1": "AR + replay",
                        "2": "AR + no replay",
                        "3": "Tablet + replay",
                        "4": "Tablet + no replay"
                        }

class MyWorkbook:
    def __init__(self):
        self.conditionCol = 3
        self.errorCol = 4
        self.correctErrorCol = 5
        self.guessedErrorCol = 6
        self.diagnoseTimeCol = 7
        #ConfidenceTimeCol = 6
        self.dateCol = 1
        self.partIdCol = 2
    
        self.file_path = rospkg.RosPack().get_path('video_logging') + '/Sheets'
        self.excel_filename = "data.xlsx"
        self.fullpath = os.path.join(self.file_path, self.excel_filename)
        self.wb = load_workbook(self.fullpath)
        self.ws = self.wb.active
    
    def write_next_row(self, participant_no, condition, error, correct_error, guessed_error, diagnosis_time):
        row = self.ws.max_row + 1
        self.ws.cell(row=row, column=self.dateCol).value = datetime.now()
        self.ws.cell(row=row, column=self.partIdCol).value = participant_no
        self.ws.cell(row=row, column=self.conditionCol).value = condition
        self.ws.cell(row=row, column=self.errorCol).value = error
        self.ws.cell(row=row, column=self.correctErrorCol).value = correct_error
        self.ws.cell(row=row, column=self.guessedErrorCol).value = guessed_error
        self.ws.cell(row=row, column=self.diagnoseTimeCol).value = diagnosis_time
        self.wb.save(self.fullpath)



class AR_error_diagnostics:
    def __init__(self, participant_no):
        self.participant_no = int(participant_no)
        self.conditions = ["","","","","","","",""]
        self.errors = ["","","","","","","",""]

    def signal_handler(self, sig, frame):
        print("\n\nshutting down\n\n")
        sys.exit(0)

    def all_loop(self):
        #Find and open file with conditions and errors used by each participant
        signal.signal(signal.SIGINT, self.signal_handler)
        pol_file_path = rospkg.RosPack().get_path('video_logging') + '/Sheets'
        pol_filename = "pol.xlsx"
        pol_wb = load_workbook(os.path.join(pol_file_path, pol_filename))
        pol_sheet = pol_wb.active
        pol_rows = 31
        pol_columns = pol_sheet.max_column
        participant_row = 0

        #Find row relating to current participant
        for i in range(2,pol_rows):
            number = int(pol_sheet.cell(row = i, column = 1).value)
            if number == self.participant_no:
                participant_row = i
        
        #Find order of conditions for current participant
        j = 0
        for i in range(2,6):
            condition = int(pol_sheet.cell(row = participant_row, column = i).value)
            self.conditions[j] = str(condition)
            j = j + 1
            self.conditions[j] = str(condition)
            j = j + 1


        #Find order of errors for current participant
        for i in range(0,8):
            col = i+6
            self.errors[i] = pol_sheet.cell(row = participant_row, column = col).value
        
        workbook = MyWorkbook()
        #Go through each error
        for l, error in enumerate(self.errors):
            #Print error information here
            print("\nThe error will be " + Errors.types[self.errors[l]])
            #Print condition for operator to know which method is being used
            print("The condition required for this error is " + Conditions.conditions[self.conditions[l]])
            skip = input("Do you want to Skip this Error/condition? (Y/N)  ")
            cont = 0
            if skip == "Y" or skip == "N":
                cont = 1
            while cont != 1:
                skip = input("Error. Invalid Input. Do you want to Skip this Error/condition? (Y/N)  ")
                if skip == "Y" or skip == "N":
                    cont = 1
            if skip == "Y":
                print("Skipping error " + Errors.types[self.errors[l]] + "\n\n\n")
                continue
            wait = input("\nPrepare condition now, then press any key to begin streaming data   ")
            print("\n")
            #Once user has confirmed ready, switch to other file to start rosbag record
            rosbag_player = rosbag_player_experimental.Play_Rosbag()
            data_to_write = rosbag_player.play_rosbag(Errors.rosbags[self.errors[l]],Errors.types[self.errors[l]])
            
            MyWorkbook().write_next_row(participant_no,Conditions.conditions[self.conditions[l]],Errors.types[self.errors[l]],data_to_write[0],data_to_write[1],data_to_write[2])
            print("Error " + str(l+1) + " completed\n\n\n")
        print("All error conditions complete, exiting...\n\n")

if __name__ == '__main__':
    rospy.init_node('error_diagnostics_user_study')
    participant_no = input("Enter Participant Number: ")
    
    AR_error_diagnostics_obj = AR_error_diagnostics(participant_no)
    AR_error_diagnostics_obj.all_loop() 
        # filename = 'participant_' + args.participant_ID + '.pkl'
        # pickle.dump(handover_obj.seq_list, open(filename, "wb"))